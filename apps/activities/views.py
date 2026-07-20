import datetime
import logging
import openpyxl
from django.db import transaction
from django.db.models import Q, Min, Max
from rest_framework import viewsets

logger = logging.getLogger(__name__)
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from apps.core.models import OrganizationalUnit, Guideline, User
from apps.core.utils import log_action, compute_diff, get_user_unit_tree
from apps.core.views import _create_notification
from apps.core.permissions import IsAdmin, HasRole, has_any_role, ROLE_EXECUTOR, ROLE_PLANNER, ROLE_APPROVER, ROLE_DIRECTOR
from apps.activities.models import Activity, ActivityGuideline, ActivityOrgUnit, ActivityMapping, UnfulfilledActivity, ActivityAttachment, ActivityComment, ActivityResponsible, ActivityParticipant
from apps.activities.serializers import (ActivitySerializer, ActivityGuidelineSerializer,
                                          ActivityOrgUnitSerializer, ActivityMappingSerializer,
                                          UnfulfilledActivitySerializer, ActivityAttachmentSerializer,
                                          ActivityCommentSerializer)
from apps.schedule.models import SchedulePeriod, SchedulePeriodMapping, ScheduleOrgUnit, ApprovedPlan
from rest_framework.exceptions import ValidationError
class ActivityViewSet(viewsets.ModelViewSet):
    queryset = Activity.objects.all().select_related(
        'category', 'organizational_unit', 'arc', 'activity_type',
        'associated_objective', 'measurement_criterion'
    ).prefetch_related('activityguideline_set')
    serializer_class = ActivitySerializer
    search_fields = ['description', 'place', 'responsible', 'participants']
    filterset_fields = ['category', 'organizational_unit', 'arc', 'activity_type', 'is_important', 'is_general']

    def get_permissions(self):
        if self.action in ['pending_approval', 'approve', 'reject', 'approve_subunit_activity', 'reject_subunit_activity', 'approve_subunit_cronograms']:
            return [HasRole(ROLE_APPROVER, ROLE_DIRECTOR)]
        if self.action in ['create', 'update', 'partial_update', 'batch_delete', 'batch_update', 'batch_assign_unit', 'assign_to_units', 'map_to_user', 'distribute_to_subunits', 'import_activities', 'destroy']:
            return [HasRole(ROLE_PLANNER, ROLE_DIRECTOR)]
        return [IsAuthenticated()]

    def get_queryset(self):
        qs = super().get_queryset()
        user = self.request.user
        if not (user.is_staff or has_any_role(user, [ROLE_DIRECTOR])):
            uos = get_user_unit_tree(user)
            mapped_act_ids = ActivityMapping.objects.filter(user=user).values_list('activity_id', flat=True)
            if has_any_role(user, [ROLE_EXECUTOR]) and not has_any_role(user, [ROLE_PLANNER, ROLE_APPROVER]):
                qs = qs.filter(id__in=mapped_act_ids)
            else:
                qs = qs.filter(
                    Q(organizational_unit__in=uos) |
                    Q(activityorgunit__organizational_unit__in=uos) |
                    Q(created_by=user) |
                    Q(responsible_user=user) |
                    Q(id__in=mapped_act_ids)
                )
        if self.action == 'list':
            desde = self.request.query_params.get('FechaDesde')
            hasta = self.request.query_params.get('FechaHasta')
            if desde:
                qs = qs.filter(schedule_periods__start_date__gte=desde)
            if hasta:
                qs = qs.filter(schedule_periods__end_date__lte=hasta)
        return qs.distinct()

    @transaction.atomic
    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)
        instance = serializer.instance
        meta = {'activity_id': instance.id, 'activity_desc': instance.description[:80]}
        for ap in instance.participant_relations.select_related('user'):
            _create_notification(ap.user, f'Se le asignó como participante en actividad: {instance.description[:80]}', 'assignment',
                                 related_object_id=instance.id, related_object_type='Activity', meta={**meta, 'user_id': ap.user.id, 'user_name': ap.user.display_name})
        for ar in instance.responsible_relations.select_related('user'):
            _create_notification(ar.user, f'Se le asignó como responsable en actividad: {instance.description[:80]}', 'assignment',
                                 related_object_id=instance.id, related_object_type='Activity', meta={**meta, 'user_id': ar.user.id, 'user_name': ar.user.display_name})
        log_action(self.request, 'Actividad', f'Creó actividad: {instance.description}')

    @transaction.atomic
    def perform_update(self, serializer):
        old = self.get_object()
        old_resp = set(old.responsible_relations.values_list('user_id', flat=True))
        old_part = set(old.participant_relations.values_list('user_id', flat=True))
        super().perform_update(serializer)
        instance = serializer.instance
        new_resp = set(instance.responsible_relations.values_list('user_id', flat=True))
        new_part = set(instance.participant_relations.values_list('user_id', flat=True))
        added_resp = new_resp - old_resp
        added_part = new_part - old_part
        if added_resp or added_part:
            from django.contrib.auth import get_user_model
            User = get_user_model()
            meta_upd = {'activity_id': instance.id, 'activity_desc': instance.description[:80]}
            for uid in added_resp:
                try:
                    u = User.objects.get(id=uid)
                    _create_notification(u, f'Se le asignó como responsable en actividad: {instance.description[:80]}', 'assignment',
                                         related_object_id=instance.id, related_object_type='Activity',
                                         meta={**meta_upd, 'user_id': u.id, 'user_name': u.display_name})
                except User.DoesNotExist:
                    pass
            for uid in added_part:
                try:
                    u = User.objects.get(id=uid)
                    _create_notification(u, f'Se le asignó como participante en actividad: {instance.description[:80]}', 'assignment',
                                         related_object_id=instance.id, related_object_type='Activity',
                                         meta={**meta_upd, 'user_id': u.id, 'user_name': u.display_name})
                except User.DoesNotExist:
                    pass
        diff = compute_diff(old, serializer.validated_data, ['description', 'place', 'responsible', 'participants', 'is_important', 'is_general'])
        log_action(self.request, 'Actividad', f'Actualizó actividad: {instance.description}', data_diff=diff)

    def perform_destroy(self, instance):
        log_action(self.request, 'Actividad', f'Eliminó actividad: {instance.description}')
        instance.delete()

    @action(detail=False, methods=['get'])
    def pending_approval(self, request):
        user = request.user
        uos = get_user_unit_tree(user)
        qs = ActivityOrgUnit.objects.filter(
            organizational_unit__in=uos,
            status__in=['', 'Re-Enviado']
        ).select_related('activity', 'organizational_unit')
        return Response(ActivityOrgUnitSerializer(qs, many=True).data)

    @action(detail=False, methods=['post'])
    def approve(self, request):
        ids = request.data.get('ids', [])
        user_uos = get_user_unit_tree(request.user)
        if request.user.is_staff or has_any_role(request.user, [ROLE_DIRECTOR]):
            user_uos = OrganizationalUnit.objects.all()
        qs = ActivityOrgUnit.objects.filter(id__in=ids, organizational_unit__in=user_uos).select_related('activity__created_by', 'organizational_unit__responsible')
        for aou in qs:
            aou.status = 'Aprobado'
            aou.save()
            ActivityMapping.objects.get_or_create(activity=aou.activity, user=request.user)
            SchedulePeriodMapping.objects.filter(
                schedule_period__activity=aou.activity,
                user=request.user
            ).delete()
            period_range = SchedulePeriod.objects.filter(activity=aou.activity).aggregate(
                min_date=Min('start_date'), max_date=Max('end_date')
            )
            sd = period_range['min_date'] or datetime.date.today()
            ed = period_range['max_date'] or datetime.date.today()
            ApprovedPlan.objects.get_or_create(
                organizational_unit=aou.organizational_unit,
                activity=aou.activity,
                defaults={'start_date': sd, 'end_date': ed, 'approved_by': request.user, 'observations': ''}
            )
            meta_approve = {'activity_id': aou.activity_id, 'activity_desc': aou.activity.description[:80]}
            for u in [aou.activity.created_by, aou.organizational_unit.responsible]:
                if u and u != request.user:
                    _create_notification(u, f'Actividad aprobada: {aou.activity.description[:80]}', 'approval',
                                         related_object_id=aou.activity_id, related_object_type='Activity',
                                         meta={**meta_approve, 'user_id': u.id, 'user_name': u.display_name})
        log_action(request, 'Actividad', f'Aprobó actividades: {ids}')
        return Response({'ok': True})

    @action(detail=False, methods=['post'])
    def reject(self, request):
        ids = request.data.get('ids', [])
        user_uos = get_user_unit_tree(request.user)
        if request.user.is_staff or has_any_role(request.user, [ROLE_DIRECTOR]):
            user_uos = OrganizationalUnit.objects.all()
        qs = ActivityOrgUnit.objects.filter(id__in=ids, organizational_unit__in=user_uos).select_related('activity__created_by', 'organizational_unit__responsible')
        for aou in qs:
            aou.status = 'Rechazado'
            aou.save()
            meta_reject = {'activity_id': aou.activity_id, 'activity_desc': aou.activity.description[:80]}
            for u in [aou.activity.created_by, aou.organizational_unit.responsible]:
                if u and u != request.user:
                    _create_notification(u, f'Actividad rechazada: {aou.activity.description[:80]}', 'rejection',
                                         related_object_id=aou.activity_id, related_object_type='Activity',
                                         meta={**meta_reject, 'user_id': u.id, 'user_name': u.display_name})
        log_action(request, 'Actividad', f'Rechazó actividades: {ids}')
        return Response({'ok': True})

    @action(detail=False, methods=['post'])
    def assign_to_units(self, request):
        activity_id = request.data.get('activity_id')
        unit_ids = request.data.get('unit_ids', [])
        if not activity_id:
            return Response({'error': 'activity_id required'}, status=400)
        for uid in unit_ids:
            ActivityOrgUnit.objects.get_or_create(
                activity_id=activity_id,
                organizational_unit_id=uid,
                defaults={'status': ''}
            )
            for sp in SchedulePeriod.objects.filter(activity_id=activity_id):
                ScheduleOrgUnit.objects.get_or_create(
                    schedule_period=sp,
                    organizational_unit_id=uid,
                    defaults={'status': ''}
                )
        try:
            activity = Activity.objects.get(id=activity_id)
        except Activity.DoesNotExist:
            return Response({'error': 'Actividad no encontrada'}, status=404)
        for uid in unit_ids:
            try:
                uo = OrganizationalUnit.objects.get(id=uid)
            except OrganizationalUnit.DoesNotExist:
                continue
            if uo.responsible:
                _create_notification(uo.responsible, f'Actividad asignada a su unidad: {activity.description[:80]}', 'assignment',
                                     related_object_id=activity_id, related_object_type='Activity',
                                     meta={'activity_id': activity_id, 'activity_desc': activity.description[:80], 'unit_id': uid, 'unit_name': uo.name})
        log_action(request, 'Actividad', f'Asignó actividad {activity_id} a UOs {unit_ids}')
        return Response({'ok': True})

    @action(detail=False, methods=['post'])
    def map_to_user(self, request):
        activity_id = request.data.get('activity_id')
        user_id = request.data.get('user_id')
        if not activity_id or not user_id:
            return Response({'error': 'activity_id and user_id required'}, status=400)
        try:
            ActivityMapping.objects.get_or_create(activity_id=activity_id, user_id=user_id)
        except Exception as e:
            return Response({'error': str(e)}, status=400)
        try:
            user = User.objects.get(id=user_id)
            activity = Activity.objects.get(id=activity_id)
            desc = (activity.description or '')[:80]
            _create_notification(user, f'Se le vinculó a actividad: {desc}', 'assignment',
                                 related_object_id=activity_id, related_object_type='Activity',
                                 meta={'activity_id': activity_id, 'activity_desc': desc, 'user_id': user.id, 'user_name': user.display_name or ''})
        except (User.DoesNotExist, Activity.DoesNotExist):
            pass
        except Exception as e:
            logger.exception('Error in map_to_user notification')
            return Response({'error': str(e)}, status=500)
        return Response({'ok': True})

    @action(detail=False, methods=['post'])
    def batch_delete(self, request):
        ids = request.data.get('ids', [])
        if not ids:
            return Response({'error': 'ids required'}, status=400)
        qs = self.get_queryset().filter(id__in=ids)
        deleted = qs.count()
        qs.delete()
        log_action(request, 'Actividad', f'Eliminación masiva: {deleted} actividades (ids: {ids})')
        return Response({'deleted': deleted})

    @action(detail=False, methods=['post'])
    def batch_update(self, request):
        ids = request.data.get('ids', [])
        updates = request.data.get('updates', {})
        if not ids:
            return Response({'error': 'ids required'}, status=400)
        allowed = {'is_important', 'is_general', 'color', 'organizational_unit', 'category', 'arc', 'activity_type', 'associated_objective', 'measurement_criterion'}
        clean = {k: v for k, v in updates.items() if k in allowed}
        if not clean:
            return Response({'error': 'no valid fields to update'}, status=400)
        qs = self.get_queryset().filter(id__in=ids)
        count = qs.count()
        qs.update(**clean)
        log_action(request, 'Actividad', f'Actualización masiva: {count} actividades con {clean}')
        return Response({'updated': count})

    @action(detail=False, methods=['post'])
    def batch_assign_unit(self, request):
        ids = request.data.get('ids', [])
        unit_id = request.data.get('organizational_unit')
        if not ids or not unit_id:
            return Response({'error': 'ids and organizational_unit required'}, status=400)
        created = 0
        for aid in ids:
            _, c = ActivityOrgUnit.objects.get_or_create(activity_id=aid, organizational_unit_id=unit_id, defaults={'status': ''})
            if c:
                created += 1
            for sp in SchedulePeriod.objects.filter(activity_id=aid):
                ScheduleOrgUnit.objects.get_or_create(
                    schedule_period=sp,
                    organizational_unit_id=unit_id,
                    defaults={'status': ''}
                )
        try:
            uo = OrganizationalUnit.objects.get(id=unit_id)
            if uo.responsible:
                _create_notification(uo.responsible, f'Actividades asignadas a su unidad ({created} items)', 'assignment',
                                     related_object_id=unit_id, related_object_type='OrganizationalUnit',
                                     meta={'unit_id': unit_id, 'unit_name': uo.name, 'count': created})
        except OrganizationalUnit.DoesNotExist:
            pass
        log_action(request, 'Actividad', f'Asignación masiva: {created} actividades a UO {unit_id}')
        return Response({'assigned': created})

    @action(detail=False, methods=['post'])
    def import_activities(self, request):
        file = request.FILES.get('file')
        if not file:
            return Response({'error': 'No file uploaded'}, status=400)
        if file.size > 10 * 1024 * 1024:
            return Response({'error': 'Archivo demasiado grande (máx 10MB)'}, status=400)
        ext = file.name.split('.')[-1].lower() if '.' in file.name else ''
        if ext not in ['xlsx', 'xls']:
            return Response({'error': 'Formato no soportado. Use archivos .xlsx o .xls'}, status=400)
        wb = openpyxl.load_workbook(file, read_only=True)
        ws = wb.active
        MAX_ROWS = 10000
        if ws.max_row and ws.max_row > MAX_ROWS:
            return Response({'error': f'El archivo excede el límite de {MAX_ROWS} filas'}, status=400)
        # Read header row to build column mapping
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), [])
        header_map = {}
        norm_headers = {
            'actividad': 'desc', 'lugar': 'place',
            'fecha': 'start_date', 'fecha inicio': 'start_date', 'fecha_inicio': 'start_date',
            'fecha fin': 'end_date', 'fecha_fin': 'end_date', 'fecha final': 'end_date', 'fecha_final': 'end_date',
            'hora': 'start_time', 'hora fin': 'end_time', 'hora_fin': 'end_time',
            'responsable': 'responsible', 'responsables': 'responsible',
            'participantes': 'participants',
            'categoria': 'category_name', 'categoría': 'category_name',
            'tipoactividad': 'type_name', 'tipo actividad': 'type_name',
            'arc': 'arc_name',
            'objetivo': 'obj_name', 'objetivo asociado': 'obj_name',
            'criteriomedida': 'crit_name', 'criterio de medida': 'crit_name',
            'uo': 'uo_name', 'unidad organizativa': 'uo_name',
            'importante': 'is_imp',
            'general': 'is_gen',
            'lineamientos': 'guideline_names', 'lineamiento': 'guideline_names',
            'responsable usuar': 'responsible_usernames', 'responsable usuario': 'responsible_usernames',
            'responsables usuar': 'responsible_usernames', 'responsables usuarios': 'responsible_usernames',
            'participante usuar': 'participant_usernames', 'participante usuario': 'participant_usernames',
            'participantes usuar': 'participant_usernames', 'participantes usuarios': 'participant_usernames',
        }
        for idx, h in enumerate(header_row):
            key = str(h).strip().lower().replace('á', 'a').replace('é', 'e').replace('í', 'i').replace('ó', 'o').replace('ú', 'u')
            if key in norm_headers:
                header_map[norm_headers[key]] = idx

        created = 0
        errors = []
        from apps.core.models import Category, ActivityType, ARC, WorkObjective, MeasurementCriterion
        for row in ws.iter_rows(min_row=2, values_only=True):
            def cell(field):
                i = header_map.get(field)
                return row[i] if i is not None and i < len(row) else None

            desc = str(cell('desc') or '').strip()
            if not desc:
                continue
            try:
                place = str(cell('place') or '').strip()
                start_date = cell('start_date')
                end_date = cell('end_date')
                start_time = cell('start_time')
                end_time = cell('end_time')
                responsible = str(cell('responsible') or '').strip()
                participants = str(cell('participants') or '').strip()
                resp_user_raw = str(cell('responsible_usernames') or '').strip()
                part_user_raw = str(cell('participant_usernames') or '').strip()
                resp_user_names = [u.strip() for u in resp_user_raw.split(';') if u.strip()] if resp_user_raw else []
                part_user_names = [u.strip() for u in part_user_raw.split(';') if u.strip()] if part_user_raw else []
                category_name = str(cell('category_name') or '').strip()
                type_name = str(cell('type_name') or '').strip()
                arc_name = str(cell('arc_name') or '').strip()
                obj_name = str(cell('obj_name') or '').strip()
                crit_name = str(cell('crit_name') or '').strip()
                uo_name = str(cell('uo_name') or '').strip()
                is_imp = str(cell('is_imp') or '').strip().lower() in ['si', 'yes', 'x', 'true', '1']
                is_gen = str(cell('is_gen') or '').strip().lower() in ['si', 'yes', 'x', 'true', '1']
                guideline_raw = str(cell('guideline_names') or '').strip()
                guideline_names = [g.strip() for g in guideline_raw.split(';') if g.strip()] if guideline_raw else []

                cat = Category.objects.filter(name=category_name).first() if category_name else None
                atype = ActivityType.objects.filter(name=type_name).first() if type_name else None
                arc_obj = ARC.objects.filter(name=arc_name).first() if arc_name else None
                obj_obj = WorkObjective.objects.filter(name=obj_name).first() if obj_name else None
                crit_obj = MeasurementCriterion.objects.filter(name=crit_name).first() if crit_name else None
                uo = OrganizationalUnit.objects.filter(name=uo_name).first() if uo_name else None
                if uo and not (self.request.user.is_staff or has_any_role(self.request.user, [ROLE_DIRECTOR])):
                    user_uos = get_user_unit_tree(self.request.user)
                    if uo not in user_uos:
                        errors.append(f'Fila {desc}: No tiene permisos para importar en la UO "{uo_name}"')
                        continue
                def parse_date(val):
                    if val is None:
                        return None
                    if isinstance(val, datetime.datetime):
                        return val.date()
                    if isinstance(val, str):
                        return datetime.datetime.strptime(val[:10], '%Y-%m-%d').date()
                    if isinstance(val, datetime.date):
                        return val
                    return None
                sd = parse_date(start_date) or datetime.date.today()
                ed = parse_date(end_date) or sd
                if sd > ed:
                    errors.append(f'Fila {desc}: La fecha fin ({ed}) es anterior a la fecha inicio ({sd})')
                    continue
                if start_time and isinstance(start_time, datetime.time):
                    st = start_time
                else:
                    st = datetime.time(8, 0)
                if end_time and isinstance(end_time, datetime.time):
                    et = end_time
                else:
                    et = datetime.time(16, 0)
                if st >= et:
                    errors.append(f'Fila {desc}: La hora fin ({et}) no es posterior a la hora inicio ({st})')
                    continue

                with transaction.atomic():
                    activity = Activity.objects.create(
                        description=desc, place=place,
                        responsible=responsible, participants=participants,
                        category=cat, activity_type=atype, arc=arc_obj,
                        associated_objective=obj_obj, measurement_criterion=crit_obj,
                        organizational_unit=uo, is_important=is_imp, is_general=is_gen,
                        created_by=request.user
                    )
                    if uo:
                        ActivityOrgUnit.objects.get_or_create(
                            activity=activity,
                            organizational_unit=uo,
                            defaults={'status': ''}
                        )
                    sp = SchedulePeriod.objects.create(
                        activity=activity,
                        start_date=sd, end_date=ed,
                        start_time=st, end_time=et,
                        status='PENDIENTE'
                    )
                    if uo:
                        ScheduleOrgUnit.objects.get_or_create(
                            schedule_period=sp,
                            organizational_unit=uo,
                            defaults={'status': ''}
                        )
                    for gn in guideline_names:
                        if gn:
                            gl = Guideline.objects.filter(name=gn).first()
                            if gl:
                                ActivityGuideline.objects.get_or_create(activity=activity, guideline=gl)
                    meta_import = {'activity_id': activity.id, 'activity_desc': activity.description[:80]}
                    for uname in resp_user_names:
                        u = User.objects.filter(Q(username=uname) | Q(display_name=uname)).first()
                        if u:
                            ActivityResponsible.objects.get_or_create(activity=activity, user=u)
                            ActivityMapping.objects.get_or_create(activity=activity, user=u)
                            _create_notification(u, f'Se le asignó como responsable en actividad importada: {activity.description[:80]}', 'assignment',
                                                 related_object_id=activity.id, related_object_type='Activity',
                                                 meta={**meta_import, 'user_id': u.id, 'user_name': u.display_name})
                    for uname in part_user_names:
                        u = User.objects.filter(Q(username=uname) | Q(display_name=uname)).first()
                        if u:
                            ActivityParticipant.objects.get_or_create(activity=activity, user=u)
                            ActivityMapping.objects.get_or_create(activity=activity, user=u)
                            _create_notification(u, f'Se le asignó como participante en actividad importada: {activity.description[:80]}', 'assignment',
                                                 related_object_id=activity.id, related_object_type='Activity',
                                                 meta={**meta_import, 'user_id': u.id, 'user_name': u.display_name})
                created += 1
            except (ValueError, TypeError, KeyError, IndexError) as e:
                errors.append(f'Fila {desc}: {str(e)}')
        return Response({'created': created, 'errors': errors})

    @action(detail=False, methods=['post'])
    def distribute_to_subunits(self, request):
        activity_id = request.data.get('activity_id')
        if not activity_id:
            return Response({'error': 'activity_id required'}, status=400)
        try:
            activity = Activity.objects.get(id=activity_id)
        except Activity.DoesNotExist:
            return Response({'error': 'Actividad no encontrada'}, status=404)
        if not activity.organizational_unit:
            return Response({'error': 'Activity has no organizational unit'}, status=400)
        sub_units = OrganizationalUnit.objects.filter(parent=activity.organizational_unit)
        for su in sub_units:
            ActivityOrgUnit.objects.get_or_create(
                activity=activity,
                organizational_unit=su,
                defaults={'status': ''}
            )
        periods = SchedulePeriod.objects.filter(activity=activity)
        for sp in periods:
            for su in sub_units:
                ScheduleOrgUnit.objects.get_or_create(
                    schedule_period=sp,
                    organizational_unit=su,
                    defaults={'status': ''}
                )
        for su in sub_units:
            if su.responsible:
                _create_notification(su.responsible, f'Actividad distribuida a su subunidad: {activity.description[:80]}', 'assignment',
                                     related_object_id=activity_id, related_object_type='Activity',
                                     meta={'activity_id': activity_id, 'activity_desc': activity.description[:80], 'unit_id': su.id, 'unit_name': su.name})
        return Response({'ok': True, 'units_distributed': sub_units.count()})

    @action(detail=False, methods=['post'])
    def approve_subunit_activity(self, request):
        aou_id = request.data.get('activity_org_unit_id')
        if not aou_id:
            return Response({'error': 'activity_org_unit_id required'}, status=400)
        user_uos = get_user_unit_tree(request.user)
        if not (request.user.is_staff or has_any_role(request.user, [ROLE_DIRECTOR])):
            qs = ActivityOrgUnit.objects.filter(id=aou_id, organizational_unit__in=user_uos)
        else:
            qs = ActivityOrgUnit.objects.filter(id=aou_id)
        try:
            aou = qs.get()
        except ActivityOrgUnit.DoesNotExist:
            return Response({'error': 'Actividad UO no encontrada o sin permisos'}, status=404)
        aou.status = 'Aprobado'
        aou.save()
        ActivityMapping.objects.get_or_create(activity=aou.activity, user=request.user)
        period_range = SchedulePeriod.objects.filter(activity=aou.activity).aggregate(
            min_date=Min('start_date'), max_date=Max('end_date')
        )
        sd = period_range['min_date'] or datetime.date.today()
        ed = period_range['max_date'] or datetime.date.today()
        ApprovedPlan.objects.get_or_create(
            organizational_unit=aou.organizational_unit,
            activity=aou.activity,
            defaults={'start_date': sd, 'end_date': ed, 'approved_by': request.user, 'observations': ''}
        )
        meta_sub = {'activity_id': aou.activity_id, 'activity_desc': aou.activity.description[:80]}
        if aou.activity.created_by and aou.activity.created_by != request.user:
            _create_notification(aou.activity.created_by, f'Actividad de subunidad aprobada: {aou.activity.description[:80]}', 'approval',
                                 related_object_id=aou.activity_id, related_object_type='Activity',
                                 meta={**meta_sub, 'user_id': aou.activity.created_by.id, 'user_name': aou.activity.created_by.display_name})
        return Response({'ok': True})

    @action(detail=False, methods=['post'])
    def reject_subunit_activity(self, request):
        aou_id = request.data.get('activity_org_unit_id')
        if not aou_id:
            return Response({'error': 'activity_org_unit_id required'}, status=400)
        user_uos = get_user_unit_tree(request.user)
        if not (request.user.is_staff or has_any_role(request.user, [ROLE_DIRECTOR])):
            qs = ActivityOrgUnit.objects.filter(id=aou_id, organizational_unit__in=user_uos)
        else:
            qs = ActivityOrgUnit.objects.filter(id=aou_id)
        try:
            aou = qs.get()
        except ActivityOrgUnit.DoesNotExist:
            return Response({'error': 'Actividad UO no encontrada o sin permisos'}, status=404)
        aou.status = 'Rechazado'
        aou.save()
        meta_rej = {'activity_id': aou.activity_id, 'activity_desc': aou.activity.description[:80]}
        if aou.activity.created_by and aou.activity.created_by != request.user:
            _create_notification(aou.activity.created_by, f'Actividad de subunidad rechazada: {aou.activity.description[:80]}', 'rejection',
                                 related_object_id=aou.activity_id, related_object_type='Activity',
                                 meta={**meta_rej, 'user_id': aou.activity.created_by.id, 'user_name': aou.activity.created_by.display_name})
        return Response({'ok': True})

    @action(detail=False, methods=['post'])
    def approve_subunit_cronograms(self, request):
        schedule_org_unit_id = request.data.get('schedule_org_unit_id')
        if not schedule_org_unit_id:
            return Response({'error': 'schedule_org_unit_id required'}, status=400)
        user_uos = get_user_unit_tree(request.user)
        if not (request.user.is_staff or has_any_role(request.user, [ROLE_DIRECTOR])):
            qs = ScheduleOrgUnit.objects.filter(id=schedule_org_unit_id, organizational_unit__in=user_uos)
        else:
            qs = ScheduleOrgUnit.objects.filter(id=schedule_org_unit_id)
        try:
            sou = qs.get()
        except ScheduleOrgUnit.DoesNotExist:
            return Response({'error': 'Cronograma UO no encontrado o sin permisos'}, status=404)
        sou.status = 'Aprobado'
        sou.save()
        SchedulePeriodMapping.objects.get_or_create(
            schedule_period=sou.schedule_period,
            user=request.user
        )
        return Response({'ok': True})


class ActivityGuidelineViewSet(viewsets.ModelViewSet):
    queryset = ActivityGuideline.objects.all()
    serializer_class = ActivityGuidelineSerializer

    def get_permissions(self):
        if self.action in ['create', 'update', 'partial_update', 'destroy']:
            return [HasRole(ROLE_PLANNER, ROLE_DIRECTOR)]
        return [IsAuthenticated()]


class ActivityOrgUnitViewSet(viewsets.ModelViewSet):
    queryset = ActivityOrgUnit.objects.all().select_related('activity', 'organizational_unit')
    serializer_class = ActivityOrgUnitSerializer
    filterset_fields = ['activity', 'organizational_unit', 'status']

    def get_permissions(self):
        if self.action in ['create', 'update', 'partial_update', 'destroy']:
            return [HasRole(ROLE_PLANNER, ROLE_DIRECTOR)]
        return [IsAuthenticated()]

    @action(detail=True, methods=['post'])
    def re_send(self, request, pk=None):
        sou = self.get_object()
        if not (request.user.is_staff or has_any_role(request.user, [ROLE_PLANNER, ROLE_DIRECTOR])):
            if has_any_role(request.user, [ROLE_APPROVER]):
                uos = get_user_unit_tree(request.user)
                if sou.organizational_unit not in uos:
                    return Response({'error': 'No tienes permiso para reenviar esta UO'}, status=403)
            else:
                return Response({'error': 'No tienes permiso para reenviar'}, status=403)
        sou.status = 'Re-Enviado'
        sou.save()
        log_action(request, 'ActividadUO', f'Re-envió actividad UO: {pk}')
        return Response({'ok': True})


class ActivityMappingViewSet(viewsets.ModelViewSet):
    queryset = ActivityMapping.objects.all()
    serializer_class = ActivityMappingSerializer

    def get_permissions(self):
        if self.action in ['create', 'update', 'partial_update', 'destroy']:
            return [HasRole(ROLE_PLANNER, ROLE_DIRECTOR)]
        return [IsAuthenticated()]


class UnfulfilledActivityViewSet(viewsets.ModelViewSet):
    queryset = UnfulfilledActivity.objects.all().select_related('registered_by', 'activity')
    serializer_class = UnfulfilledActivitySerializer
    filterset_fields = ['activity', 'schedule_period']

    def get_permissions(self):
        if self.action in ['create']:
            return [HasRole(ROLE_PLANNER, ROLE_DIRECTOR, ROLE_EXECUTOR)]
        if self.action in ['update', 'partial_update', 'destroy']:
            return [HasRole(ROLE_PLANNER, ROLE_DIRECTOR)]
        return [IsAuthenticated()]

    def perform_create(self, serializer):
        serializer.save(registered_by=self.request.user)
        log_action(self.request, 'ActividadIncumplida', 'Registró actividad incumplida')

    def perform_update(self, serializer):
        super().perform_update(serializer)
        log_action(self.request, 'ActividadIncumplida', 'Actualizó actividad incumplida')

    def perform_destroy(self, instance):
        log_action(self.request, 'ActividadIncumplida', 'Eliminó actividad incumplida')
        instance.delete()


ALLOWED_ATTACHMENT_EXTENSIONS = None  # uses settings.ALLOWED_UPLOAD_EXTENSIONS

def _validate_file_extension(file):
    from django.conf import settings
    allowed = settings.ALLOWED_UPLOAD_EXTENSIONS
    ext = file.name.rsplit('.', 1)[-1].lower() if '.' in file.name else ''
    if ext not in allowed:
        raise ValidationError({'error': f'Extensión de archivo no permitida: .{ext}'})
    if file.size > 10 * 1024 * 1024:
        raise ValidationError({'error': 'Archivo demasiado grande (máx 10MB)'})

class ActivityAttachmentViewSet(viewsets.ModelViewSet):
    queryset = ActivityAttachment.objects.all().select_related('uploaded_by')
    serializer_class = ActivityAttachmentSerializer
    filterset_fields = ['activity']
    ordering = ['id']

    def get_permissions(self):
        if self.action in ['create']:
            return [HasRole(ROLE_EXECUTOR, ROLE_PLANNER, ROLE_APPROVER, ROLE_DIRECTOR)]
        if self.action in ['update', 'partial_update', 'destroy']:
            return [HasRole(ROLE_PLANNER, ROLE_DIRECTOR)]
        return [IsAuthenticated()]

    def perform_create(self, serializer):
        file = serializer.validated_data.get('file')
        if file:
            _validate_file_extension(file)
        serializer.save(uploaded_by=self.request.user)
        log_action(self.request, 'AdjuntoActividad', 'Adjuntó archivo a actividad')

    def perform_destroy(self, instance):
        log_action(self.request, 'AdjuntoActividad', 'Eliminó adjunto de actividad')
        instance.delete()

    @action(detail=True, methods=['get'])
    def download(self, request, pk=None):
        attachment = self.get_object()
        if not attachment.file:
            return Response({'error': 'Archivo no encontrado'}, status=404)
        from django.http import FileResponse
        response = FileResponse(attachment.file.open('rb'), as_attachment=True, filename=attachment.file.name.split('/')[-1])
        return response


class ActivityCommentViewSet(viewsets.ModelViewSet):
    queryset = ActivityComment.objects.all().select_related('user')
    serializer_class = ActivityCommentSerializer
    filterset_fields = ['activity']
    ordering = ['id']

    def get_permissions(self):
        if self.action in ['create']:
            return [HasRole(ROLE_EXECUTOR, ROLE_PLANNER, ROLE_APPROVER, ROLE_DIRECTOR)]
        if self.action in ['update', 'partial_update', 'destroy']:
            return [HasRole(ROLE_PLANNER, ROLE_DIRECTOR)]
        return [IsAuthenticated()]

    def perform_create(self, serializer):
        serializer.save(user=self.request.user)
        log_action(self.request, 'ComentarioActividad', 'Comentó en actividad')

    def perform_destroy(self, instance):
        log_action(self.request, 'ComentarioActividad', 'Eliminó comentario de actividad')
        instance.delete()
