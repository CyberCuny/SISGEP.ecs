import datetime
import io
import openpyxl
from django.db.models import Q, Count
from django.http import HttpResponse
from rest_framework import viewsets, status
from rest_framework.exceptions import ValidationError
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from apps.core.models import User, OrganizationalUnit
from apps.core.utils import log_action, compute_diff
from apps.core.views import _create_notification
from apps.activities.models import ActivityMapping
from apps.schedule.models import (SchedulePeriod, SchedulePeriodMapping,
                                  ScheduleOrgUnit, WorkDay, ApprovedPlan, ScheduleComment)
from apps.schedule.serializers import (SchedulePeriodSerializer, SchedulePeriodMappingSerializer,
                                        ScheduleOrgUnitSerializer, WorkDaySerializer,
                                        ApprovedPlanSerializer, ScheduleCommentSerializer)
from apps.core.throttles import ReportThrottle


def _check_overlap(activity_id, start_date, end_date, start_time, end_time, exclude_id=None):
    qs = SchedulePeriod.objects.filter(
        activity_id=activity_id,
        start_date__lte=end_date,
        end_date__gte=start_date,
        start_time__lt=end_time,
        end_time__gt=start_time,
    )
    if exclude_id:
        qs = qs.exclude(id=exclude_id)
    return qs.exists()


class SchedulePeriodViewSet(viewsets.ModelViewSet):
    queryset = SchedulePeriod.objects.all().select_related('activity')
    serializer_class = SchedulePeriodSerializer
    search_fields = ['description', 'observation']
    filterset_fields = ['activity', 'status', 'start_date', 'end_date', 'is_extraplan', 'has_incidence']

    def get_queryset(self):
        qs = super().get_queryset()
        user = self.request.user
        if not user.is_staff:
            uos = OrganizationalUnit.objects.filter(responsible=user)
            mapped_period_ids = SchedulePeriodMapping.objects.filter(user=user).values_list('schedule_period_id', flat=True)
            mapped_act_ids = ActivityMapping.objects.filter(user=user).values_list('activity_id', flat=True)
            qs = qs.filter(
                Q(activity__organizational_unit__in=uos) |
                Q(id__in=mapped_period_ids) |
                Q(activity_id__in=mapped_act_ids)
            )
        return qs.distinct()

    def perform_create(self, serializer):
        data = serializer.validated_data
        if _check_overlap(data['activity'].id, data['start_date'], data['end_date'], data['start_time'], data['end_time']):
            raise ValidationError({'error': 'Solapamiento detectado'})
        super().perform_create(serializer)
        log_action(self.request, 'CronogramaPeriodo', f'Creó periodo para actividad {data["activity"].id}')

    def perform_update(self, serializer):
        instance = self.get_object()
        data = serializer.validated_data
        if _check_overlap(data.get('activity', instance.activity).id,
                          data.get('start_date', instance.start_date),
                          data.get('end_date', instance.end_date),
                          data.get('start_time', instance.start_time),
                          data.get('end_time', instance.end_time),
                          exclude_id=instance.id):
            raise ValidationError({'error': 'Solapamiento detectado'})
        super().perform_update(serializer)
        diff = compute_diff(instance, serializer.validated_data, ['start_date', 'end_date', 'start_time', 'end_time', 'status', 'description', 'observation'])
        log_action(self.request, 'CronogramaPeriodo', f'Actualizó periodo {instance.id}', object_id=instance.pk, data_diff=diff)

    def perform_destroy(self, instance):
        user = self.request.user
        activity = instance.activity
        if user.is_staff or activity.organizational_unit and activity.organizational_unit.responsible == user:
            log_action(self.request, 'CronogramaPeriodo', f'Eliminó periodo {instance.id}')
            instance.delete()
        else:
            SchedulePeriodMapping.objects.filter(schedule_period=instance, user=user).delete()
            log_action(self.request, 'CronogramaPeriodo', f'Removió referencia a periodo {instance.id}')

    @action(detail=False, methods=['get'])
    def calendar(self, request):
        try:
            year = int(request.query_params.get('year', datetime.date.today().year))
            month = int(request.query_params.get('month', datetime.date.today().month))
        except (ValueError, TypeError):
            return Response({'error': 'year y month deben ser números enteros'}, status=400)
        first_day = datetime.date(year, month, 1)
        last_day = datetime.date(year + 1, 1, 1) - datetime.timedelta(days=1) if month == 12 else datetime.date(year, month + 1, 1) - datetime.timedelta(days=1)
        qs = self.get_queryset().filter(
            start_date__lte=last_day,
            end_date__gte=first_day
        )
        activity_id = request.query_params.get('activity_id')
        status_filter = request.query_params.get('status')
        org_unit_id = request.query_params.get('org_unit_id')
        user_id = request.query_params.get('user_id')
        if activity_id:
            qs = qs.filter(activity_id=activity_id)
        if status_filter:
            qs = qs.filter(status=status_filter)
        if org_unit_id:
            qs = qs.filter(activity__organizational_unit_id=org_unit_id)
        if user_id:
            try:
                u = User.objects.get(id=user_id)
                uos = OrganizationalUnit.objects.filter(responsible=u)
                mapped_period_ids = SchedulePeriodMapping.objects.filter(user=u).values_list('schedule_period_id', flat=True)
                mapped_act_ids = ActivityMapping.objects.filter(user=u).values_list('activity_id', flat=True)
                qs = qs.filter(
                    Q(activity__organizational_unit__in=uos) |
                    Q(id__in=mapped_period_ids) |
                    Q(activity_id__in=mapped_act_ids)
                )
            except User.DoesNotExist:
                pass
        data = []
        for sp in qs:
            data.append({
                'id': sp.id,
                'title': sp.activity.description or 'Sin desc',
                'start_date': sp.start_date.isoformat(),
                'end_date': sp.end_date.isoformat(),
                'start_time': sp.start_time.isoformat() if sp.start_time else None,
                'end_time': sp.end_time.isoformat() if sp.end_time else None,
                'color': sp.color or sp.activity.color,
                'status': sp.status,
                'is_extraplan': sp.is_extraplan,
                'has_incidence': sp.has_incidence,
                'is_modified': sp.is_modified,
                'observation': sp.observation,
                'activity_id': sp.activity_id,
                'activity_description': sp.activity.description,
            })
        return Response(data)

    @action(detail=False, methods=['get'])
    def individual_calendar(self, request):
        try:
            year = int(request.query_params.get('year', datetime.date.today().year))
            month = int(request.query_params.get('month', datetime.date.today().month))
        except (ValueError, TypeError):
            return Response({'error': 'year y month deben ser números enteros'}, status=400)
        first_day = datetime.date(year, month, 1)
        last_day = datetime.date(year + 1, 1, 1) - datetime.timedelta(days=1) if month == 12 else datetime.date(year, month + 1, 1) - datetime.timedelta(days=1)
        qs = self.get_queryset().filter(
            start_date__lte=last_day,
            end_date__gte=first_day
        )
        user_id = request.query_params.get('user_id')
        if user_id:
            try:
                u = User.objects.get(id=user_id)
                uos = OrganizationalUnit.objects.filter(responsible=u)
                mapped_period_ids = SchedulePeriodMapping.objects.filter(user=u).values_list('schedule_period_id', flat=True)
                mapped_act_ids = ActivityMapping.objects.filter(user=u).values_list('activity_id', flat=True)
                qs = qs.filter(
                    Q(activity__organizational_unit__in=uos) |
                    Q(id__in=mapped_period_ids) |
                    Q(activity_id__in=mapped_act_ids)
                )
            except User.DoesNotExist:
                pass
        return Response(SchedulePeriodSerializer(qs, many=True).data)

    @action(detail=False, methods=['post'])
    def update_status(self, request):
        ids = request.data.get('ids', [])
        status_val = request.data.get('status', '')
        allowed = {'CUMPLIDO', 'INCUMPLIDO', 'PENDIENTE', ''}
        if status_val not in allowed:
            return Response({'error': f'Estado inválido: {status_val}'}, status=400)
        SchedulePeriod.objects.filter(id__in=ids).update(status=status_val)
        log_action(request, 'CronogramaPeriodo', f'Actualizó estado de periodos {ids} a {status_val}')
        return Response({'ok': True})

    @action(detail=True, methods=['patch'])
    def update_single_status(self, request, pk=None):
        instance = self.get_object()
        instance.status = request.data.get('status', instance.status)
        instance.observation = request.data.get('observation', instance.observation)
        instance.save()
        return Response(SchedulePeriodSerializer(instance).data)

    @action(detail=True, methods=['patch'])
    def drag_drop(self, request, pk=None):
        instance = self.get_object()
        from datetime import datetime
        start_date = request.data.get('start_date')
        end_date = request.data.get('end_date')
        if start_date:
            try:
                datetime.strptime(start_date, '%Y-%m-%d').date()
            except (ValueError, TypeError):
                return Response({'error': 'Formato de fecha inválido (use YYYY-MM-DD)'}, status=400)
        if end_date:
            try:
                datetime.strptime(end_date, '%Y-%m-%d').date()
            except (ValueError, TypeError):
                return Response({'error': 'Formato de fecha inválido (use YYYY-MM-DD)'}, status=400)
        start_date = start_date or instance.start_date
        end_date = end_date or instance.end_date
        if _check_overlap(instance.activity_id, start_date, end_date, instance.start_time, instance.end_time, exclude_id=instance.id):
            return Response({'error': 'Solapamiento detectado'}, status=status.HTTP_400_BAD_REQUEST)
        instance.start_date = start_date
        instance.end_date = end_date
        instance.is_modified = True
        instance.save()
        log_action(request, 'CronogramaPeriodo', f'Reagendó periodo {instance.id}')
        return Response(SchedulePeriodSerializer(instance).data)

    @action(detail=False, methods=['get'])
    def incidences(self, request):
        desde = request.query_params.get('desde')
        hasta = request.query_params.get('hasta')
        qs = self.get_queryset().filter(has_incidence=True)
        if desde:
            qs = qs.filter(start_date__gte=desde)
        if hasta:
            qs = qs.filter(end_date__lte=hasta)
        return Response(SchedulePeriodSerializer(qs, many=True).data)

    @action(detail=False, methods=['get'])
    def compliance_stats(self, request):
        desde = request.query_params.get('desde')
        hasta = request.query_params.get('hasta')
        user_id = request.query_params.get('user_id')
        qs = self.get_queryset()
        if desde:
            qs = qs.filter(start_date__gte=desde)
        if hasta:
            qs = qs.filter(end_date__lte=hasta)
        if user_id:
            try:
                u = User.objects.get(id=user_id)
                uos = OrganizationalUnit.objects.filter(responsible=u)
                qs = qs.filter(activity__organizational_unit__in=uos)
            except User.DoesNotExist:
                pass

        total = qs.count()
        cumplidas = qs.filter(status='CUMPLIDO').count()
        incumplidas = qs.filter(status='INCUMPLIDO').count()
        pendientes = qs.filter(Q(status='PENDIENTE') | Q(status__isnull=True)).count()
        extraplan = qs.filter(is_extraplan=True, status='CUMPLIDO').count()
        modificadas = qs.filter(is_modified=True).count()

        def pct(val):
            if total == 0:
                return 0
            return round(val / total * 100, 1)

        try:
            page = int(request.query_params.get('page', 1))
            page_size = int(request.query_params.get('page_size', 20))
        except (ValueError, TypeError):
            return Response({'error': 'page y page_size deben ser números enteros'}, status=400)
        page_size = min(page_size, 500)
        offset = (page - 1) * page_size
        detail_qs = qs.select_related('activity').order_by('start_date')[offset:offset + page_size]
        detalle = []
        for sp in detail_qs:
            detalle.append({
                'id': sp.id,
                'activity': sp.activity.description,
                'activity_id': sp.activity_id,
                'plan_type': 'Extraplan' if sp.is_extraplan else 'Normal',
                'start_date': sp.start_date.isoformat(),
                'end_date': sp.end_date.isoformat(),
                'status': sp.status or 'PENDIENTE',
                'observation': sp.observation or '',
                'has_incidence': sp.has_incidence,
                'is_extraplan': sp.is_extraplan,
                'is_modified': sp.is_modified,
            })

        monthly_data = {}
        if desde and hasta:
            try:
                year = int(desde[:4])
            except (ValueError, TypeError):
                return Response({'error': 'desde debe ser una fecha válida (YYYY-MM-DD)'}, status=400)
            for m in range(1, 13):
                m_start = datetime.date(year, m, 1)
                m_end = datetime.date(year + (1 if m == 12 else 0), (1 if m == 12 else m + 1), 1) - datetime.timedelta(days=1)
                m_total = qs.filter(start_date__lte=m_end, end_date__gte=m_start).count()
                monthly_data[f'month_{m:02d}'] = m_total

        return Response({
            'total': total,
            'cumplidas': cumplidas,
            'incumplidas': incumplidas,
            'pendientes': pendientes,
            'extraplan': extraplan,
            'modificadas': modificadas,
            'pct_cumplidas': pct(cumplidas),
            'pct_incumplidas': pct(incumplidas),
            'pct_pendientes': pct(pendientes),
            'pct_extraplan': pct(extraplan),
            'detalle_count': len(detalle),
            'detalle_total': total,
            'detalle': detalle,
            **monthly_data,
        })

    @action(detail=False, methods=['get'])
    def annual_calendar(self, request):
        try:
            year = int(request.query_params.get('year', datetime.date.today().year))
        except (ValueError, TypeError):
            return Response({'error': 'year debe ser un número entero'}, status=400)
        first_day = datetime.date(year, 1, 1)
        last_day = datetime.date(year, 12, 31)
        qs = self.get_queryset().filter(
            start_date__lte=last_day,
            end_date__gte=first_day
        ).select_related('activity__category')
        activity_id = request.query_params.get('activity_id')
        status_filter = request.query_params.get('status')
        org_unit_id = request.query_params.get('org_unit_id')
        user_id = request.query_params.get('user_id')
        if activity_id:
            qs = qs.filter(activity_id=activity_id)
        if status_filter:
            qs = qs.filter(status=status_filter)
        if org_unit_id:
            qs = qs.filter(activity__organizational_unit_id=org_unit_id)
        if user_id:
            try:
                u = User.objects.get(id=user_id)
                uos = OrganizationalUnit.objects.filter(responsible=u)
                mapped_period_ids = SchedulePeriodMapping.objects.filter(user=u).values_list('schedule_period_id', flat=True)
                mapped_act_ids = ActivityMapping.objects.filter(user=u).values_list('activity_id', flat=True)
                qs = qs.filter(
                    Q(activity__organizational_unit__in=uos) |
                    Q(id__in=mapped_period_ids) |
                    Q(activity_id__in=mapped_act_ids)
                )
            except User.DoesNotExist:
                pass

        months = {m: [] for m in range(1, 13)}
        for sp in qs:
            start_m = sp.start_date.month if sp.start_date.year == year else 1
            end_m = sp.end_date.month if sp.end_date.year == year else 12
            for m in range(start_m, end_m + 1):
                if m in months:
                    months[m].append({
                        'id': sp.id,
                        'activity': sp.activity.description,
                        'activity_id': sp.activity_id,
                        'category': sp.activity.category.name if sp.activity.category else 'Sin categoría',
                        'category_id': sp.activity.category_id,
                        'color': sp.color or sp.activity.color or '#3b82f6',
                        'start_date': sp.start_date.isoformat(),
                        'end_date': sp.end_date.isoformat(),
                        'status': sp.status,
                    })

        categories = {}
        for m, items in months.items():
            for item in items:
                cat = item['category']
                if cat not in categories:
                    categories[cat] = {}
                if m not in categories[cat]:
                    categories[cat][m] = []
                categories[cat][m].append(item)

        return Response({
            'year': year,
            'months': months,
            'categories': categories,
        })

    @action(detail=False, methods=['get'])
    def pending_cronograms_approval(self, request):
        user = request.user
        uos = OrganizationalUnit.objects.filter(responsible=user)
        qs = ScheduleOrgUnit.objects.filter(
            organizational_unit__in=uos,
            status__in=['', 'Re-Enviado']
        ).select_related('schedule_period__activity', 'organizational_unit')
        data = []
        for sou in qs:
            data.append({
                'id': sou.id,
                'schedule_period_id': sou.schedule_period_id,
                'activity_description': sou.schedule_period.activity.description,
                'organizational_unit_name': sou.organizational_unit.name,
                'status': sou.status,
                'start_date': sou.schedule_period.start_date,
                'end_date': sou.schedule_period.end_date,
            })
        return Response(data)

    @action(detail=False, methods=['post'])
    def approve_cronogram_org_unit(self, request):
        ids = request.data.get('ids', [])
        ScheduleOrgUnit.objects.filter(id__in=ids).update(status='Aprobado')
        for sou in ScheduleOrgUnit.objects.filter(id__in=ids).select_related('schedule_period__activity', 'organizational_unit__responsible'):
            SchedulePeriodMapping.objects.get_or_create(
                schedule_period=sou.schedule_period,
                user=request.user
            )
            for u in [sou.schedule_period.activity.created_by, sou.organizational_unit.responsible]:
                if u and u != request.user:
                    _create_notification(u, f'Cronograma aprobado: {sou.schedule_period.activity.description[:80]}', 'approval',
                                         related_object_id=sou.schedule_period_id, related_object_type='SchedulePeriod')
        return Response({'ok': True})

    @action(detail=False, methods=['post'])
    def reject_cronogram_org_unit(self, request):
        ids = request.data.get('ids', [])
        for sou in ScheduleOrgUnit.objects.filter(id__in=ids).select_related('schedule_period__activity', 'organizational_unit__responsible'):
            sou.status = 'Rechazado'
            sou.save()
            for u in [sou.schedule_period.activity.created_by, sou.organizational_unit.responsible]:
                if u and u != request.user:
                    _create_notification(u, f'Cronograma rechazado: {sou.schedule_period.activity.description[:80]}', 'rejection',
                                         related_object_id=sou.schedule_period_id, related_object_type='SchedulePeriod')
        log_action(request, 'Cronograma', f'Rechazó cronogramas: {ids}')
        return Response({'ok': True})


class SchedulePeriodMappingViewSet(viewsets.ModelViewSet):
    queryset = SchedulePeriodMapping.objects.all()
    serializer_class = SchedulePeriodMappingSerializer


class ScheduleOrgUnitViewSet(viewsets.ModelViewSet):
    queryset = ScheduleOrgUnit.objects.all().select_related('organizational_unit')
    serializer_class = ScheduleOrgUnitSerializer
    filterset_fields = ['schedule_period', 'organizational_unit', 'status']
    search_fields = ['organizational_unit__name']

    @action(detail=True, methods=['post'])
    def re_send(self, request, pk=None):
        sou = self.get_object()
        sou.status = 'Re-Enviado'
        sou.save()
        log_action(request, 'CronogramaUO', f'Re-envió cronograma UO: {pk}')
        return Response({'ok': True})


class WorkDayViewSet(viewsets.ModelViewSet):
    queryset = WorkDay.objects.all()
    serializer_class = WorkDaySerializer
    filterset_fields = ['user']

    @action(detail=False, methods=['post'])
    def save_batch(self, request):
        user_id = request.data.get('user_id')
        days = request.data.get('days', [])
        if not user_id:
            return Response({'error': 'user_id required'}, status=400)
        WorkDay.objects.filter(user_id=user_id).delete()
        for d in days:
            WorkDay.objects.create(user_id=user_id, day=d)
        return Response({'ok': True})


class ApprovedPlanViewSet(viewsets.ModelViewSet):
    queryset = ApprovedPlan.objects.all()
    serializer_class = ApprovedPlanSerializer
    filterset_fields = ['organizational_unit']
    ordering = ['id']



class ReportsViewSet(viewsets.ViewSet):
    permission_classes = [IsAuthenticated]
    throttle_classes = [ReportThrottle]

    @action(detail=False, methods=['get'])
    def individual(self, request):
        user_id = request.query_params.get('user_id')
        periods = SchedulePeriod.objects.all().select_related('activity')
        if user_id:
            try:
                u = User.objects.get(id=user_id)
                uos = OrganizationalUnit.objects.filter(responsible=u)
                mapped = SchedulePeriodMapping.objects.filter(user_id=user_id).values_list('schedule_period_id', flat=True)
                periods = periods.filter(
                    Q(activity__organizational_unit__in=uos) | Q(id__in=mapped)
                )
            except User.DoesNotExist:
                pass

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Plan Individual'
        ws.append(['Actividad', 'Fecha Inicio', 'Fecha Fin', 'Hora Inicio', 'Hora Fin', 'Estado', 'Observación'])
        for p in periods:
            ws.append([
                p.activity.description, str(p.start_date), str(p.end_date),
                str(p.start_time or ''), str(p.end_time or ''), p.status or '', p.observation or ''
            ])
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=plan_individual.xlsx'
        wb.save(response)
        return response

    @action(detail=False, methods=['get'])
    def export_ics(self, request):
        desde = request.query_params.get('desde')
        hasta = request.query_params.get('hasta')
        periods = SchedulePeriod.objects.all().select_related('activity')
        if desde:
            periods = periods.filter(start_date__gte=desde)
        if hasta:
            periods = periods.filter(end_date__lte=hasta)

        lines = [
            'BEGIN:VCALENDAR',
            'VERSION:2.0',
            'PRODID:-//PlanTrabajo//ES',
            'CALSCALE:GREGORIAN',
            'METHOD:PUBLISH',
        ]
        for p in periods:
            uid = f'plan-{p.id}@plantrabajo'
            dtstart = p.start_date.strftime('%Y%m%d') + 'T' + (p.start_time.strftime('%H%M%S') if p.start_time else '000000')
            dtend = p.end_date.strftime('%Y%m%d') + 'T' + (p.end_time.strftime('%H%M%S') if p.end_time else '235959')
            desc = (p.activity.description or '').replace(',', '\\,').replace(';', '\\;')
            lines.extend([
                'BEGIN:VEVENT',
                f'UID:{uid}',
                f'DTSTART:{dtstart}',
                f'DTEND:{dtend}',
                f'SUMMARY:{desc}',
                f'DESCRIPTION:{(p.description or "").replace(",", "\\,").replace(";", "\\;")}',
                'END:VEVENT',
            ])
        lines.append('END:VCALENDAR')
        response = HttpResponse('\r\n'.join(lines), content_type='text/calendar; charset=utf-8')
        response['Content-Disposition'] = 'attachment; filename=calendar.ics'
        return response

    @action(detail=False, methods=['get'])
    def import_template(self, request):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Plantilla Importación'
        headers = ['Actividad', 'Lugar', 'Fecha', 'Hora', 'Hora Fin', 'Responsable', 'Participantes',
                   'Categoría', 'TipoActividad', 'ARC', 'Objetivo', 'CriterioMedida',
                   'UO', 'Importante', 'General', 'Lineamientos']
        ws.append(headers)
        ws.append(['Ejemplo de actividad', 'Oficina', '2026-01-15', '08:00', '17:00', 'Juan Pérez', 'María López',
                   'Capacitación', 'Taller', 'ARC 1', 'Mejorar procesos', 'Criterio A',
                   'Dirección General', 'Sí', '', 'Lineamiento 1; Lineamiento 2'])
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 18
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=plantilla_importacion.xlsx'
        wb.save(response)
        return response

    @action(detail=False, methods=['get'])
    def compliance_pdf(self, request):
        desde = request.query_params.get('desde')
        hasta = request.query_params.get('hasta')
        periods = SchedulePeriod.objects.all()
        if desde:
            periods = periods.filter(start_date__gte=desde)
        if hasta:
            periods = periods.filter(end_date__lte=hasta)

        total = periods.count()
        cumplidas = periods.filter(status='CUMPLIDO').count()
        incumplidas = periods.filter(status='INCUMPLIDO').count()
        pendientes = periods.filter(Q(status='PENDIENTE') | Q(status__isnull=True)).count()
        pct = round(cumplidas / total * 100, 1) if total else 0

        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        from reportlab.lib import colors
        from reportlab.lib.units import mm

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=20*mm, bottomMargin=20*mm)
        styles = getSampleStyleSheet()
        styles.add(ParagraphStyle('Title2', parent=styles['Title'], fontSize=18, spaceAfter=12))
        styles.add(ParagraphStyle('Normal2', parent=styles['Normal'], fontSize=11, spaceAfter=6))

        elements = []
        elements.append(Paragraph('Reporte de Cumplimiento', styles['Title2']))
        elements.append(Paragraph(f'Periodo: {desde or "inicio"} - {hasta or "hoy"}', styles['Normal2']))
        elements.append(Spacer(1, 12))

        data = [
            ['Indicador', 'Valor'],
            ['Total actividades', str(total)],
            ['Cumplidas', f'{cumplidas} ({pct}%)'],
            ['Incumplidas', f'{incumplidas} ({round(incumplidas/total*100,1) if total else 0}%)'],
            ['Pendientes', f'{pendientes} ({round(pendientes/total*100,1) if total else 0}%)'],
        ]
        t = Table(data, colWidths=[200, 200])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3b82f6')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 11),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f1f5f9')]),
        ]))
        elements.append(t)
        elements.append(Spacer(1, 20))
        elements.append(Paragraph(f'Generado: {datetime.datetime.now().strftime("%Y-%m-%d %H:%M")}', styles['Normal2']))

        doc.build(elements)
        buffer.seek(0)
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename=reporte_cumplimiento.pdf'
        return response

    @action(detail=False, methods=['get'])
    def by_uo(self, request):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Cumplimiento por UO'
        ws.append(['U. Organizativa', 'Total', 'Cumplidas', 'Incumplidas', 'Pendientes', '% Cumplimiento'])
        uos = OrganizationalUnit.objects.annotate(
            total=Count('activity_set__schedule_periods', distinct=True),
            cumplidas=Count('activity_set__schedule_periods',
                filter=Q(activity_set__schedule_periods__status='CUMPLIDO'), distinct=True),
            incumplidas=Count('activity_set__schedule_periods',
                filter=Q(activity_set__schedule_periods__status='INCUMPLIDO'), distinct=True),
            pendientes=Count('activity_set__schedule_periods',
                filter=Q(activity_set__schedule_periods__status__isnull=True) |
                       Q(activity_set__schedule_periods__status='PENDIENTE'), distinct=True),
        )
        for uo in uos:
            pct = round(uo.cumplidas / uo.total * 100, 1) if uo.total else 0
            ws.append([uo.name, uo.total, uo.cumplidas, uo.incumplidas, uo.pendientes, pct])
        for col in range(1, 7):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=reporte_por_uo.xlsx'
        wb.save(response)
        return response

    @action(detail=False, methods=['get'])
    def comparative(self, request):
        try:
            year = int(request.query_params.get('year', str(datetime.date.today().year)))
        except (ValueError, TypeError):
            return Response({'error': 'year debe ser un número entero'}, status=400)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Comparativo Mensual'
        ws.append(['Mes', 'Total', 'Cumplidas', 'Incumplidas', 'Pendientes', '% Cumplimiento'])
        for month in range(1, 13):
            periods = SchedulePeriod.objects.filter(start_date__year=int(year), start_date__month=month)
            total = periods.count()
            cumplidas = periods.filter(status='CUMPLIDO').count()
            incumplidas = periods.filter(status='INCUMPLIDO').count()
            pendientes = periods.filter(status__isnull=True).count()
            pct = round(cumplidas / total * 100, 1) if total else 0
            month_name = ['Enero','Febrero','Marzo','Abril','Mayo','Junio','Julio','Agosto','Septiembre','Octubre','Noviembre','Diciembre'][month-1]
            ws.append([month_name, total, cumplidas, incumplidas, pendientes, pct])
        for col in range(1, 7):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=reporte_comparativo.xlsx'
        wb.save(response)
        return response


class ScheduleCommentViewSet(viewsets.ModelViewSet):
    queryset = ScheduleComment.objects.all().select_related('user')
    serializer_class = ScheduleCommentSerializer
    filterset_fields = ['schedule_period']

    def perform_create(self, serializer):
        serializer.save(user=self.request.user)
        log_action(self.request, 'ComentarioCronograma', 'Comentó en cronograma')

    def perform_destroy(self, instance):
        log_action(self.request, 'ComentarioCronograma', 'Eliminó comentario de cronograma')
        instance.delete()
